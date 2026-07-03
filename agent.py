"""
Hotel Mirror — Lease Intelligence Agent
----------------------------------------
Searches European hotel & building lease markets for serviced apartment operators.
No scoring. No priorities. Just clean, structured data:
property details · landlord contacts · broker contacts · source links.

Output: JSON + CSV ready for import into Notion / Airtable / Google Sheets.

Markets: DE · NL · FR · BE · LU · CH · CZ · GB · IE · IT · ES · PT · GR · PL
"""

import os
import json
import csv
import re
import datetime
import time
import argparse

# Load .env file if present (requires python-dotenv, installed via requirements.txt)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed — rely on env vars set in shell

import anthropic

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

COUNTRIES = {
    "AT": "Austria", "BE": "Belgium", "NL": "Netherlands", "LU": "Luxembourg",
    "CZ": "Czech Republic", "HU": "Hungary", "PL": "Poland",
    "FR": "France", "DE": "Germany",
    "ES": "Spain", "PT": "Portugal", "IT": "Italy",
    "AE": "UAE", "IL": "Israel",
    "DK": "Denmark", "SE": "Sweden", "FI": "Finland", "NO": "Norway",
    "CH": "Switzerland", "GB": "United Kingdom", "IE": "Ireland",
}

# Target markets — key expansion cities.
# Searches are restricted to these target cities only.
TARGET_CITIES = {
    "AT": ["Graz", "Innsbruck", "Kitzbühel", "Salzburg", "Vienna"],
    "BE": ["Antwerp", "Bruges", "Brussels", "Ghent"],
    "NL": ["Amsterdam", "Groningen", "Haarlem", "Maastricht", "Rotterdam", "The Hague", "Utrecht"],
    "LU": ["Luxembourg"],
    "CZ": ["Prague"],
    "HU": ["Budapest"],
    "PL": ["Krakow", "Warsaw"],
    "FR": ["Aix-en-Provence", "Bordeaux", "Cannes", "Lille", "Lyon", "Marseille",
           "Montpellier", "Nantes", "Nice", "Paris", "Strasbourg", "Toulouse"],
    "DE": ["Berlin", "Bonn", "Bremen", "Cologne", "Dresden", "Düsseldorf", "Frankfurt",
           "Freiburg im Breisgau", "Hamburg", "Heidelberg", "Leipzig", "Munich",
           "Nuremberg", "Stuttgart"],
    "ES": ["Alicante", "Barcelona", "Bilbao", "Granada", "Las Palmas de Gran Canaria",
           "Madrid", "Malaga", "Marbella", "Palma", "San Sebastian", "Sevilla", "Valencia"],
    "PT": ["Funchal", "Lisbon", "Porto"],
    "IT": ["Bologna", "Florence", "Genoa", "Milano", "Napoli", "Palermo", "Rome",
           "Turin", "Venice", "Verona"],
    "AE": ["Dubai"],
    "IL": ["Tel Aviv"],
    "DK": ["Copenhagen"],
    "SE": ["Gothenburg", "Stockholm"],
    "FI": ["Helsinki"],
    "NO": ["Oslo"],
    "CH": ["Basel", "Bern", "Geneva", "Lausanne", "Lugano", "Luzern", "Zurich"],
    "GB": ["Bath", "Belfast", "Birmingham", "Brighton", "Bristol", "Cardiff",
           "Edinburgh", "Glasgow", "Leeds", "Liverpool",
           "London", "Manchester", "Newcastle", "Oxford", "York"],
    "IE": ["Dublin"],
}

BROKERS = (
    "Christie & Co, JLL Hotels & Hospitality, CBRE Hotels, Savills Hotels, "
    "BNP Paribas Real Estate, Colliers Hotels, Knight Frank Hotels, "
    "HCRE, Hospitality Advisors, Cushman & Wakefield, Streimer, "
    "Immobilienscout24, Funda, SeLoger, Immowelt, Rightmove Commercial, "
    "Idealista, Immobiliare.it"
)

SYSTEM_PROMPT = f"""You are a real estate researcher finding ACTIVE, VERIFIABLE listings of hotels
and buildings available for LEASE (not purchase) to operators across Europe.
Client: a serviced apartment / extended-stay operator that leases, never buys.

CRITICAL SEARCH STRATEGY — public lease listings exist under LOCAL-LANGUAGE terms:
- Germany/Austria/Switzerland: "Hotel pachten", "Hotel zu verpachten", "Hotelpacht",
  "Boardinghouse pachten" — immobilienscout24.de (Gewerbe), immowelt.de, hotel-boerse.com
- France: "hotel a louer", "location-gerance hotel", "murs et fonds hotel",
  "fonds de commerce hotel" — SeLoger Bureaux & Commerces, CessionPME, leboncoin pro
- Italy: "hotel in affitto", "albergo in gestione" — immobiliare.it, idealista.it
- Spain/Portugal: "hotel en alquiler", "traspaso hotel", "hotel para arrendar" — idealista, fotocasa
- Netherlands/Belgium: "hotel te huur", "hotel ter overname" — funda in business, horecamakelaardij.nl
- UK/Ireland: "hotel to lease", "leasehold hotel for sale", "hotel to let" — christie.com,
  daltonsbusiness.com, rightmove commercial, businessesforsale.com
- Czech/Poland/Greece: "hotel k pronajmu", "hotel do wynajecia" — sreality.cz, otodom.pl, spitogatos.gr

ALSO INCLUDE:
- Leasehold hotel businesses for sale (the lease itself is being sold)
- Location-gerance and Pacht offers (both are lease structures)
- Operator/tenant searches announced by landlords or brokers in hospitality news
- New developments seeking operators (forward lease deals)

VALIDATION RULES:
- Every result MUST come from an actual search result — include its real source_url
- Never invent listings, contacts, or URLs. Use null for unknown fields — partial data is fine
- Prefer listings published or updated 2024-2026
- Contact details only if actually shown in the listing

TARGET MARKET RULE:
Include properties located in the target cities from the user message OR their metro
areas (suburbs and adjacent municipalities count). If after thorough searching you have
fewer than 5 in-target results, you may add strong listings from elsewhere in the SAME
target countries — mark these with "notes": "outside target city list".
NEVER return an empty array if you found any qualifying listings.

SEARCH PROCEDURE — depth over breadth:
Run at least 6 distinct, CITY-SPECIFIC searches using local-language terms, prioritising
markets with the most public lease supply (Germany, Italy, Spain, France, Netherlands)
among the selected targets. Examples: "Hotel pachten Berlin", "Hotel zu verpachten
Muenchen", "hotel in affitto Milano", "albergo in gestione Roma", "traspaso hotel Madrid",
"hotel location-gerance Paris", "hotel te huur Amsterdam". Do NOT run one generic
Europe-wide search.

Return ONLY a valid JSON array — no markdown fences, no preamble.
Return ALL validated listings you find — no minimum, up to 100 maximum.
If output space runs low, STOP adding listings and close the JSON array properly —
a complete array of 15 beats a truncated array of 40. Keep "notes" under 20 words.
Each object must have exactly these keys (use null if unknown):

{{
  "property_name":      string,
  "address":            string,
  "city":               string,
  "country_code":       "AT"|"BE"|"NL"|"LU"|"CZ"|"HU"|"PL"|"FR"|"DE"|"ES"|"PT"|"IT"|"AE"|"IL"|"DK"|"SE"|"FI"|"NO"|"CH"|"GB"|"IE",
  "property_type":      "Hotel"|"Apart-hotel"|"Office conversion"|"Residential building"|"Mixed-use",
  "rooms":              string,
  "sqm":                string,
  "floors":             string,
  "year_built":         string,
  "lease_term":         string,
  "asking_rent":        string,
  "available_from":     string,
  "landlord_name":      string,
  "landlord_company":   string,
  "landlord_email":     string,
  "landlord_phone":     string,
  "landlord_website":   string,
  "broker_name":        string,
  "broker_company":     string,
  "broker_email":       string,
  "broker_phone":       string,
  "source_url":         string,
  "source_name":        string,
  "notes":              string
}}
"""

# ─────────────────────────────────────────────
# API CLIENTS — Anthropic primary, OpenAI fallback
# ─────────────────────────────────────────────

_anthropic_key = os.environ.get("ANTHROPIC_API_KEY", "")
_openai_key = os.environ.get("OPENAI_API_KEY", "")

if not _anthropic_key and not _openai_key:
    raise SystemExit(
        "\n  ✗  No API key set.\n\n"
        "  Set at least one in your shell or .env file:\n"
        "    ANTHROPIC_API_KEY=sk-ant-...   (primary)\n"
        "    OPENAI_API_KEY=sk-...          (fallback)\n\n"
        "  Get keys at: https://console.anthropic.com / https://platform.openai.com\n"
    )

client = anthropic.Anthropic(api_key=_anthropic_key) if _anthropic_key else None


def _search_openai(user_message: str) -> str:
    """Fallback: OpenAI Responses API with built-in web_search tool."""
    import urllib.request

    body = json.dumps({
        "model": "gpt-4.1",
        "tools": [{"type": "web_search"}],
        "instructions": SYSTEM_PROMPT,
        "input": user_message,
        "max_output_tokens": 30000,
    }).encode()

    req = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=body,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {_openai_key}",
        },
    )
    with urllib.request.urlopen(req, timeout=180) as resp:
        data = json.loads(resp.read())

    text = ""
    for item in data.get("output", []):
        if item.get("type") == "message":
            for c in item.get("content", []):
                if c.get("type") == "output_text":
                    text += c.get("text", "")
    return text

# ─────────────────────────────────────────────
# AGENTIC SEARCH LOOP
# ─────────────────────────────────────────────

def search(user_message: str, max_iterations: int = 12) -> str:
    """
    Anthropic primary, OpenAI fallback (no credit, auth error, outage).
    web_search_20250305 is a SERVER-SIDE tool: Anthropic executes searches itself
    and returns results inline. Long turns pause with stop_reason 'pause_turn';
    we resume by appending the assistant content and re-calling.
    """
    if client is None:
        print("  ⚠  No Anthropic key — using OpenAI directly.")
        return _search_openai(user_message)

    try:
        return _search_anthropic(user_message, max_iterations)
    except Exception as exc:
        if not _openai_key:
            raise
        print(f"  ⚠  Anthropic failed ({exc}). Falling back to OpenAI…")
        return _search_openai(user_message)


def _search_anthropic(user_message: str, max_iterations: int = 12) -> str:
    messages = [{"role": "user", "content": user_message}]
    tools = [{"type": "web_search_20250305", "name": "web_search", "max_uses": 10}]

    for i in range(max_iterations):
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=20000,
            system=SYSTEM_PROMPT,
            tools=tools,
            messages=messages,
        )

        # Server-side search paused mid-turn — resume with full content
        if response.stop_reason == "pause_turn":
            messages.append({"role": "assistant", "content": response.content})
            print(f"  …  search round {i + 1}")
            continue

        # Done (end_turn, max_tokens, etc.) — extract all text blocks
        return "".join(
            b.text for b in response.content if b.type == "text"
        )

    return ""


# ─────────────────────────────────────────────
# PARSE JSON FROM RESPONSE
# ─────────────────────────────────────────────

def parse_listings(raw: str) -> list[dict]:
    raw = re.sub(r"```json\s*", "", raw)
    raw = re.sub(r"```\s*", "", raw).strip()
    s = raw.find("[")
    if s == -1:
        return []

    # Attempt 1: clean parse
    e = raw.rfind("]")
    if e > s:
        try:
            return json.loads(raw[s : e + 1])
        except json.JSONDecodeError:
            pass

    # Attempt 2: repair truncated output — cut to last complete object, close array
    body = raw[s:]
    last_obj = body.rfind("}")
    if last_obj == -1:
        return []
    try:
        recovered = json.loads(body[: last_obj + 1] + "]")
        print(f"  ⚠  Recovered {len(recovered)} listings from truncated JSON")
        return recovered
    except json.JSONDecodeError as err:
        print(f"  ⚠  JSON parse error (repair failed): {err}")
        return []


# ─────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────

FIELD_ORDER = [
    "property_name", "address", "city", "country_code", "property_type",
    "rooms", "sqm", "floors", "year_built",
    "lease_term", "asking_rent", "available_from",
    "landlord_name", "landlord_company", "landlord_email", "landlord_phone", "landlord_website",
    "broker_name", "broker_company", "broker_email", "broker_phone",
    "source_url", "source_name", "notes",
    "run_timestamp",
]

def stamp(listings: list[dict]) -> list[dict]:
    ts = datetime.datetime.utcnow().isoformat()
    for l in listings:
        l["run_timestamp"] = ts
        # Ensure all fields exist
        for f in FIELD_ORDER:
            l.setdefault(f, None)
    return listings

def save_json(listings: list[dict], path: str):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(listings, f, indent=2, ensure_ascii=False)
    print(f"  💾  JSON  →  {path}")

def save_csv(listings: list[dict], path: str):
    if not listings:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELD_ORDER, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(listings)
    print(f"  💾  CSV   →  {path}")



def save_xlsx(listings: list[dict], path: str):
    if not listings:
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  ⚠  openpyxl not installed — skipping .xlsx (pip install openpyxl)")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"
    headers = [f.replace("_", " ").title() for f in FIELD_ORDER]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1C1914")
        cell.alignment = Alignment(vertical="center")
    for l in listings:
        ws.append([l.get(f) for f in FIELD_ORDER])
    for col_idx, f in enumerate(FIELD_ORDER, 1):
        width = max(len(headers[col_idx - 1]),
                    *(len(str(l.get(f) or "")) for l in listings))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(50, width + 2)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"  💾  XLSX  →  {path}")


# ─────────────────────────────────────────────
# OPTIONAL: SLACK ALERT
# ─────────────────────────────────────────────

def slack_alert(listing: dict, webhook: str):
    try:
        import urllib.request
        text = (
            f"🏨 *{listing.get('property_name','?')}* — "
            f"{listing.get('city','?')}, {listing.get('country_code','?')}\n"
            f"Type: {listing.get('property_type','?')} · "
            f"Rooms: {listing.get('rooms','?')} · "
            f"Rent: {listing.get('asking_rent','?')}\n"
            f"Landlord: {listing.get('landlord_name','?')} "
            f"({listing.get('landlord_company','?')}) "
            f"{listing.get('landlord_email') or listing.get('landlord_phone','')}\n"
            f"Broker: {listing.get('broker_name','?')} "
            f"({listing.get('broker_company','?')}) "
            f"{listing.get('broker_email') or listing.get('broker_phone','')}\n"
            f"Source: {listing.get('source_url','?')}"
        )
        body = json.dumps({"text": text}).encode()
        req = urllib.request.Request(
            webhook, data=body, headers={"Content-Type": "application/json"}
        )
        urllib.request.urlopen(req, timeout=5)
    except Exception as exc:
        print(f"  ⚠  Slack alert failed: {exc}")


# ─────────────────────────────────────────────
# MAIN RUNNER
# ─────────────────────────────────────────────

def run(
    countries: list[str] | None = None,
    property_type: str = "",
    min_units: int = 0,
    city: str = "",
    notes: str = "",
    slack_webhook: str = "",
    output_dir: str = ".",
    max_results: int = 100,
) -> list[dict]:

    countries = countries or list(COUNTRIES.keys())
    country_str = ", ".join(COUNTRIES[c] for c in countries if c in COUNTRIES)

    print()
    print("══════════════════════════════════════════════════")
    print("  HOTEL MIRROR  —  Lease Intelligence Agent")
    print("══════════════════════════════════════════════════")
    print(f"  Markets  : {country_str}")
    if property_type: print(f"  Type     : {property_type}")
    if min_units:     print(f"  Min units: {min_units}")
    if city:          print(f"  City     : {city}")
    if notes:         print(f"  Notes    : {notes}")
    print()

    # Build search query
    filters = []
    if property_type: filters.append(f"property type: {property_type}")
    if min_units:     filters.append(f"minimum {min_units} units/rooms")
    if city:          filters.append(f"city or region: {city}")
    if notes:         filters.append(notes)

    target_markets = "; ".join(
        f"{COUNTRIES[c]} ({', '.join(TARGET_CITIES[c])})" for c in countries if c in COUNTRIES
    )
    query = (
        f"Find hotels and buildings currently available for lease to operators "
        f"in these target markets: {target_markets}. "
        + (f"Filters: {'; '.join(filters)}. " if filters else "")
        + "Search the web NOW using local-language lease terms per country "
        "(e.g. 'Hotel zu verpachten' for Germany, 'hotel location-gerance' for France, "
        "'hotel in affitto' for Italy, 'hotel en alquiler' for Spain). Run multiple searches "
        "across property portals, business-transfer sites, and hospitality news. Include "
        "leasehold sales and Pacht/location-gerance offers. Run at least 6 city-specific "
        "searches, depth over breadth. Return ALL validated results "
        f"(up to {max_results}) as a JSON array — never empty if you found qualifying listings."
    )

    print("  🔍  Searching the web…")
    raw = search(query)

    listings = parse_listings(raw)
    if not listings:
        print("  ⚠  No listings found. Try broader criteria.")
        return []

    listings = stamp(listings)
    print(f"  ✅  {len(listings)} listing(s) found.\n")

    # Print summary table to terminal
    print(f"  {'PROPERTY':<35} {'CITY':<18} {'CC':<4} {'ROOMS':<8} {'RENT'}")
    print(f"  {'─'*35} {'─'*18} {'─'*4} {'─'*8} {'─'*20}")
    for l in listings:
        print(
            f"  {str(l.get('property_name','?'))[:34]:<35} "
            f"{str(l.get('city','?'))[:17]:<18} "
            f"{str(l.get('country_code','?')):<4} "
            f"{str(l.get('rooms','?'))[:7]:<8} "
            f"{str(l.get('asking_rent','?'))[:25]}"
        )

    # Slack alerts
    if slack_webhook:
        print(f"\n  📣  Sending {len(listings)} Slack alert(s)…")
        for l in listings:
            slack_alert(l, slack_webhook)
            time.sleep(0.3)

    # Export
    ts = datetime.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    os.makedirs(output_dir, exist_ok=True)
    print()
    save_json(listings, f"{output_dir}/hotel_mirror_{ts}.json")
    save_csv(listings,  f"{output_dir}/hotel_mirror_{ts}.csv")
    save_xlsx(listings, f"{output_dir}/hotel_mirror_{ts}.xlsx")
    print()
    print("  ✅  Done.\n")

    return listings


# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Hotel Mirror — lease intelligence agent for serviced apartment operators",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python agent.py
  python agent.py --countries DE NL FR --type Hotel --min-units 80
  python agent.py --city Frankfurt --notes "15+ year lease preferred"
  python agent.py --countries GB IE --slack https://hooks.slack.com/services/...
        """,
    )
    parser.add_argument(
        "--countries", nargs="+", default=list(COUNTRIES.keys()),
        metavar="CC",
        help=f"ISO country codes (default: all). Options: {' '.join(COUNTRIES.keys())}",
    )
    parser.add_argument("--type",      default="", metavar="TYPE",    help="Property type filter")
    parser.add_argument("--min-units", default=0,  type=int,          help="Minimum unit/room count")
    parser.add_argument("--city",      default="", metavar="CITY",    help="City or region focus")
    parser.add_argument("--notes",     default="", metavar="TEXT",    help="Additional search criteria")
    parser.add_argument("--max",       default=100, type=int,         help="Max results (default 100)")
    parser.add_argument("--slack",     default="", metavar="URL",     help="Slack webhook URL")
    parser.add_argument("--output",    default=".", metavar="DIR",    help="Output directory (default: .)")

    args = parser.parse_args()

    run(
        countries=args.countries,
        property_type=args.type,
        min_units=args.min_units,
        city=args.city,
        notes=args.notes,
        slack_webhook=args.slack,
        output_dir=args.output,
        max_results=args.max,
    )
